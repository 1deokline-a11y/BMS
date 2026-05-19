"""엑셀 BOM 파일 파서 - 한국 광통신장비 BOM 포맷 지원"""
import os
import re
from typing import Optional
import openpyxl
from openpyxl.utils import get_column_letter


HEADER_KEYWORDS = {
    "part_number": ["품번", "부품번호", "자재번호", "part no", "part number", "구매품번"],
    "part_name": ["품명", "부품명", "자재명", "part name", "명칭"],
    "spec": ["규격", "사양", "specification", "spec"],
    "unit": ["단위", "unit"],
    "quantity": ["수량", "qty", "quantity", "사용량"],
    "notes": ["비고", "remark", "note", "참고"],
}

SKIP_VALUES = {"", None, "-", "N/A", "n/a"}


def _normalize(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _find_header_row(ws) -> tuple[Optional[int], dict]:
    """BOM 헤더 행 인덱스와 컬럼 매핑 반환"""
    for row_idx in range(1, min(20, ws.max_row + 1)):
        row_vals = [_normalize(ws.cell(row=row_idx, column=c).value).lower()
                    for c in range(1, ws.max_column + 1)]
        col_map = {}
        for field, keywords in HEADER_KEYWORDS.items():
            for col_i, cell_val in enumerate(row_vals, start=1):
                if any(kw in cell_val for kw in keywords) and field not in col_map:
                    col_map[field] = col_i
                    break
        # 품번+품명+수량 중 2개 이상 매칭되면 헤더로 인정
        matched = sum(1 for f in ["part_number", "part_name", "quantity"] if f in col_map)
        if matched >= 2:
            return row_idx, col_map
    return None, {}


def _parse_metadata_from_filename(filename: str) -> dict:
    """파일명에서 메타데이터 파싱
    예: A11-D1-0040_27형_144C_오스트리아, NGB향.xlsx
    """
    base = os.path.splitext(os.path.basename(filename))[0]
    parts = base.split("_")

    part_number = parts[0] if parts else base
    m = re.match(r"([A-Z]\d{2}-[A-Z]\d)-(\d{4})", part_number)
    product_group = m.group(1) if m else part_number[:6]
    variant_code = m.group(2) if m else "0000"

    name = "_".join(parts[1:]) if len(parts) > 1 else ""
    country_spec = ""
    spec = ""
    customer = ""

    for p in parts[1:]:
        if "향" in p or "austria" in p.lower():
            country_spec = p
        elif re.search(r"\d+C", p):
            spec = p
        elif re.search(r"[A-Z]{2,}", p) and len(p) <= 15:
            customer = p

    return {
        "part_number": part_number,
        "product_group": product_group,
        "variant_code": variant_code,
        "name": name,
        "customer": customer,
        "country_spec": country_spec,
        "spec": spec,
    }


def parse_excel_bom(filepath: str) -> dict:
    """엑셀 BOM 파일을 파싱하여 제품 정보와 BOM 항목 반환"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    meta = _parse_metadata_from_filename(filepath)
    bom_items = []

    # BOM 시트 탐색 (첫 번째 시트 또는 'BOM' 포함 시트 우선)
    target_sheet = wb.worksheets[0]
    for ws in wb.worksheets:
        if "bom" in ws.title.lower() or "부품" in ws.title:
            target_sheet = ws
            break

    ws = target_sheet
    header_row, col_map = _find_header_row(ws)
    if header_row is None:
        return {"meta": meta, "bom_items": []}

    # 셀 병합 해제용 값 가져오기
    merged_ranges = [str(mr) for mr in ws.merged_cells.ranges]

    def get_cell_value(row, col):
        cell = ws.cell(row=row, column=col)
        val = cell.value
        # 병합 셀인 경우 마스터 셀 값 찾기
        if val is None:
            for mr_str in merged_ranges:
                try:
                    mr = openpyxl.worksheet.cell_range.CellRange(mr_str)
                    if (mr.min_row <= row <= mr.max_row and
                            mr.min_col <= col <= mr.max_col):
                        val = ws.cell(row=mr.min_row, column=mr.min_col).value
                        break
                except Exception:
                    pass
        return _normalize(val)

    row_order = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        part_number = get_cell_value(row_idx, col_map.get("part_number", 1))
        part_name = get_cell_value(row_idx, col_map.get("part_name", 2))

        if part_number in SKIP_VALUES and part_name in SKIP_VALUES:
            continue
        if not part_number and not part_name:
            continue

        qty_raw = get_cell_value(row_idx, col_map.get("quantity", 0)) if "quantity" in col_map else "1"
        try:
            quantity = float(qty_raw) if qty_raw else 1.0
        except ValueError:
            quantity = 1.0

        bom_items.append({
            "part_number": part_number or f"UNKNOWN-{row_order}",
            "part_name": part_name or "",
            "spec": get_cell_value(row_idx, col_map.get("spec", 0)) if "spec" in col_map else "",
            "unit": get_cell_value(row_idx, col_map.get("unit", 0)) if "unit" in col_map else "EA",
            "quantity": quantity,
            "notes": get_cell_value(row_idx, col_map.get("notes", 0)) if "notes" in col_map else "",
            "row_order": row_order,
        })
        row_order += 1

    return {"meta": meta, "bom_items": bom_items}
