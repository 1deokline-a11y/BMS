"""BOM 데이터를 엑셀로 내보내기"""
import io
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# 색상 정의
COLOR_ADDED = "C6EFCE"       # 초록
COLOR_REMOVED = "FFC7CE"     # 빨강
COLOR_QTY_DIFF = "FFEB9C"   # 노랑
COLOR_SPEC_DIFF = "FCE4D6"  # 주황
COLOR_HEADER = "4472C4"     # 헤더 파란색
COLOR_WHITE = "FFFFFF"
COLOR_LIGHT_GRAY = "F2F2F2"

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _apply_fill(cell, color: str):
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


def _header_style(cell, text: str):
    cell.value = text
    cell.font = Font(bold=True, color=COLOR_WHITE)
    _apply_fill(cell, COLOR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER


def export_bom_to_excel(product: dict, bom_items: list) -> bytes:
    """단일 BOM을 엑셀 파일로 내보내기"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"

    headers = ["순번", "품번", "품명", "규격", "단위", "수량", "비고"]
    widths = [6, 25, 30, 30, 8, 8, 20]
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        _header_style(ws.cell(row=1, column=col), h)
        ws.column_dimensions[get_column_letter(col)].width = w

    for i, item in enumerate(bom_items, start=1):
        part = item.get("part", item)
        row = i + 1
        fill = PatternFill(start_color=COLOR_LIGHT_GRAY if i % 2 == 0 else COLOR_WHITE,
                           end_color=COLOR_LIGHT_GRAY if i % 2 == 0 else COLOR_WHITE,
                           fill_type="solid")
        values = [
            i,
            part.get("part_number", ""),
            part.get("part_name", ""),
            part.get("spec", ""),
            part.get("unit", "EA"),
            item.get("quantity", 1),
            item.get("notes", ""),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_comparison_to_excel(product1: dict, product2: dict,
                                items1: list, items2: list,
                                diff_result: dict) -> bytes:
    """BOM 비교 결과를 엑셀로 내보내기"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM 비교"

    p1_num = product1.get("part_number", "BOM1")
    p2_num = product2.get("part_number", "BOM2")

    # 제목 행
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"BOM 비교: {p1_num} vs {p2_num}  ({datetime.now().strftime('%Y-%m-%d %H:%M')})"
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center")

    # 헤더
    headers_left = ["품번", "품명", "규격", "수량"]
    headers_right = ["품번", "품명", "규격", "수량"]
    col_labels = (
        [f"[{p1_num}] {h}" for h in headers_left] +
        [f"[{p2_num}] {h}" for h in headers_right] +
        ["구분"]
    )
    for col, label in enumerate(col_labels, start=1):
        _header_style(ws.cell(row=2, column=col), label)
    ws.column_dimensions["I"].width = 12

    diff_map = {d["key"]: d for d in diff_result.get("diffs", [])}
    all_keys = list(dict.fromkeys(
        [f"{i['part']['part_number']}|{i['part']['part_name']}" for i in items1] +
        [f"{i['part']['part_number']}|{i['part']['part_name']}" for i in items2]
    ))

    idx1 = {f"{i['part']['part_number']}|{i['part']['part_name']}": i for i in items1}
    idx2 = {f"{i['part']['part_number']}|{i['part']['part_name']}": i for i in items2}

    row = 3
    for key in all_keys:
        d = diff_map.get(key, {})
        status = d.get("status", "same")
        color = {
            "added": COLOR_ADDED,
            "removed": COLOR_REMOVED,
            "qty_diff": COLOR_QTY_DIFF,
            "spec_diff": COLOR_SPEC_DIFF,
        }.get(status, COLOR_WHITE)

        i1 = idx1.get(key)
        i2 = idx2.get(key)

        def part_vals(item):
            if not item:
                return ["", "", "", ""]
            p = item.get("part", item)
            return [p.get("part_number", ""), p.get("part_name", ""),
                    p.get("spec", ""), item.get("quantity", "")]

        vals = part_vals(i1) + part_vals(i2) + [status]
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            _apply_fill(cell, color)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
        row += 1

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
